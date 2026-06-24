"""One-shot generator for data_validations.xlsx — a simplified inventory
of what the app checks per track and how each problem can be fixed.

Run: python data_validations.xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROWS = [
    # (check, what it catches, where you see it, how to fix it, fix quality)
    ("Very short distance",
     "Track is under 200 m total. Usually an accidental start.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "Delete the activity, or Approve to silence the flag",
     "No real fix"),
    ("Very short duration",
     "Track is under 60 seconds long.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "Delete the activity, or Approve to silence the flag",
     "No real fix"),
    ("Impossibly fast top speed",
     "Recorded max speed over 150 km/h.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "If GPS spikes are also flagged, run Repair on that activity. Otherwise Trim out the bad section or Approve.",
     "Auto-fix only when it's caused by a spike"),
    ("Implausible climb rate",
     "More than 200 m of climbing per kilometre of distance.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "Trim out the bad section on the activity page, or Approve",
     "No auto-fix"),
    ("GPS teleport",
     "Track jumps more than 1 km of distance in under 5 minutes.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "Trim around the jump on the activity page, or Approve",
     "No auto-fix"),
    ("GPS speed spikes",
     "Short cluster of points with impossibly fast speeds (typically Strava-export artefacts).",
     "Review page → GPS Spikes tab (map, speed charts, before/after preview); also Logs page flag",
     "Repair (clamps the bad legs and re-derives stats), or Approve",
     "Full coverage"),
    ("GPS jitter",
     "Recorded path is more than 30 % longer than a smoothed version — distance inflated by noise.",
     "Logs page — yellow flag on the row; 'Show only flagged' filter",
     "Approve only (the distance number stays inflated)",
     "No auto-fix"),
    ("Likely duplicate rides",
     "Two or more activities on the same date with distance and duration within 5 % of each other.",
     "Review page → Duplicates tab",
     "Delete one side, Exclude one side from stats, or mark 'Not duplicates' to hide the pair",
     "Full coverage"),
    ("Night-time start",
     "Activity starts between 9 PM and 7 AM local time — usually a timezone bug, sometimes a real night ride.",
     "Review page → Odd Times tab",
     "No fix offered — diagnostic only. Open the ride to investigate.",
     "Diagnostic only"),
    ("Missing activity type",
     "Activity has no type (MTB / hike / ski / etc.) and doesn't match any region's default.",
     "Review page → Missing Types tab (with a thumbnail map per row)",
     "Click a type chip to assign one, or add a Region for that area so future rides get tagged automatically",
     "Full coverage"),
]

HEADERS = ["Check", "What it catches", "Where you see it", "How to fix it", "Fix quality"]

FILL_HEADER  = PatternFill("solid", fgColor="1F2937")
FILL_GOOD    = PatternFill("solid", fgColor="DCFCE7")
FILL_PARTIAL = PatternFill("solid", fgColor="FEF9C3")
FILL_NONE    = PatternFill("solid", fgColor="FECACA")
FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
FONT_BODY    = Font(size=11)
ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
ALIGN_HEAD   = Alignment(wrap_text=True, vertical="center", horizontal="left")
BORDER       = Border(*[Side(style="thin", color="D1D5DB")] * 4)

def fill_for(quality: str) -> PatternFill:
    if quality.startswith("Full"):       return FILL_GOOD
    if quality.startswith("Diagnostic"): return FILL_PARTIAL
    if quality.startswith("Auto-fix"):   return FILL_PARTIAL
    return FILL_NONE

wb = Workbook()
ws = wb.active
ws.title = "Validations"

# Header row
for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_HEAD; c.border = BORDER

# Data rows
for r_idx, row in enumerate(ROWS, start=2):
    for c_idx, val in enumerate(row, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.font = FONT_BODY; c.alignment = ALIGN_WRAP; c.border = BORDER
    # Tint the last column by fix quality so good / partial / none read at a glance
    ws.cell(row=r_idx, column=5).fill = fill_for(row[4])

# Column widths (chars) — chosen so each cell renders ~3 lines of wrap at most
widths = [22, 50, 50, 60, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row heights — first row is the header, then each data row gets enough
# height to show wrapped text without truncating.
ws.row_dimensions[1].height = 28
for r_idx in range(2, len(ROWS) + 2):
    ws.row_dimensions[r_idx].height = 58

ws.freeze_panes = "A2"

# Second sheet: gap summary so the file works as a one-page reference.
ws2 = wb.create_sheet("Coverage summary")
ws2["A1"] = "Status"
ws2["B1"] = "Checks"
for c in (ws2["A1"], ws2["B1"]):
    c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_HEAD; c.border = BORDER

summary = [
    ("Full coverage (detect + auto-fix)", [r[0] for r in ROWS if r[4] == "Full coverage"]),
    ("Auto-fix only when caused by a spike", [r[0] for r in ROWS if r[4].startswith("Auto-fix")]),
    ("Diagnostic only — no fix offered",     [r[0] for r in ROWS if r[4].startswith("Diagnostic")]),
    ("Detect only — no auto-fix",            [r[0] for r in ROWS if r[4] == "No auto-fix" or r[4] == "No real fix"]),
]
r = 2
for label, items in summary:
    a = ws2.cell(row=r, column=1, value=label); a.font = FONT_BODY; a.alignment = ALIGN_WRAP; a.border = BORDER
    b = ws2.cell(row=r, column=2, value=", ".join(items) if items else "—"); b.font = FONT_BODY; b.alignment = ALIGN_WRAP; b.border = BORDER
    a.fill = fill_for(label.split()[0] if "Full" in label else label[:9])
    r += 1
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 80
for i in range(2, r):
    ws2.row_dimensions[i].height = 40

OUT = "data_validations.xlsx"
wb.save(OUT)
print(f"wrote {OUT}")
